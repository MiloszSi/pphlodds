"""
Turns the weekly PPHL mastersheets into the data blob the playoff-odds page
runs on, and writes it straight into the HTML file.

Usage (local files):
    python generate_pphl_data.py ^
        --premier "PPHL Premier Season 2 Mastersheet.xlsx" ^
        --challenger "Copy of PPHL Challenger Season 2 Mastersheet.xlsx" ^
        --html pphl_playoff_odds_full.html

Usage (straight from Google Sheets -- no manual "Download as .xlsx" step,
which is what caused the data to go stale before. Works as long as the
sheets are shared as "anyone with the link can view"):
    python generate_pphl_data.py ^
        --premier "https://docs.google.com/spreadsheets/d/XXXX/edit" ^
        --challenger "https://docs.google.com/spreadsheets/d/YYYY/edit" ^
        --html pphl_playoff_odds_full.html

--premier/--challenger accept either a local .xlsx path or a Google Sheets
URL (detected automatically) -- mix and match freely.

What it reads from each mastersheet (StandingsSchedule / Backend Schedule /
Backend Players tabs -- all on the "Backend" sheets so formatting on the
public-facing tabs can't break the parse):
    - Standings: rank, W/OTW/OTL/L, points, GD/GF/GA, and division membership
      (division sub-tables are detected automatically -- Premier has none,
      Challenger has Interior/Coastal).
    - Schedule: every game is "completed" if it has a score, "remaining" if
      not. OT rate is computed from the completed games.
    - Roster: goals / goals-against / takeaways / giveaways are summed per
      team from the individual player rows.

Team colors, short display names, and how many teams make the playoffs are
NOT in the spreadsheet, so those live in team_meta.json next to this script.
Add a team there the first time it appears; the script will warn (not fail)
and use a placeholder if you forget.
"""
import argparse
import json
import re
import sys
import urllib.request
import urllib.error
from tempfile import NamedTemporaryFile
from datetime import date
from pathlib import Path

import openpyxl

FALLBACK_PALETTE = [
    "#8E8E93", "#5AC8FA", "#FF9F0A", "#BF5AF2", "#32ADE6", "#FF375F",
]

GOOGLE_SHEETS_ID_RE = re.compile(r"docs\.google\.com/spreadsheets/d/([a-zA-Z0-9_-]+)")


def resolve_xlsx(source, label):
    """If `source` is a Google Sheets URL, downloads the whole workbook as .xlsx to a
    temp file and returns that path. Otherwise assumes it's already a local .xlsx path
    and returns it unchanged. Requires the sheet be shared as "anyone with the link
    can view" -- a private sheet will fail here with a clear error rather than
    silently downloading a Google login page."""
    m = GOOGLE_SHEETS_ID_RE.search(source)
    if not m:
        return source

    sheet_id = m.group(1)
    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    print(f"  Downloading {label} from Google Sheets ({sheet_id})...")
    try:
        with urllib.request.urlopen(export_url, timeout=30) as resp:
            data = resp.read()
    except urllib.error.HTTPError as e:
        raise SystemExit(f"Could not download the {label} sheet ({e.code} {e.reason}). "
                          f"Make sure it's shared as 'anyone with the link can view'.")

    if data[:2] != b"PK":  # .xlsx files are zip archives; anything else is an error/login page
        raise SystemExit(f"The {label} sheet did not download as a valid .xlsx file -- "
                          f"it's probably not shared publicly ('anyone with the link can view').")

    tmp = NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(data)
    tmp.close()
    return tmp.name

STANDINGS_COLS = {"rank": 18, "team": 19, "w": 20, "otw": 21, "otl": 22,
                   "l": 23, "points": 24, "gd": 25, "gf": 26, "ga": 27}
SCHED_COLS = {"date": 2, "away": 3, "home": 4, "away_score": 5,
              "home_score": 6, "ot": 7}
ROSTER_COLS = {"team": 2, "g": 10, "tk": 20, "gv": 21, "ga": 28}


def parse_standings(ws):
    """Returns (standings_list, divisions_dict_or_None)."""
    title_rows = []
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=STANDINGS_COLS["rank"]).value
        if isinstance(v, str) and "Standings" in v:
            title_rows.append((row, v.strip()))

    if not title_rows:
        raise ValueError("Could not find a '... Standings' title in StandingsSchedule column R")

    def read_table(title_row):
        data_start = title_row + 4
        rows = []
        r = data_start
        while r <= ws.max_row:
            team = ws.cell(row=r, column=STANDINGS_COLS["team"]).value
            if not team:
                break
            get = lambda key: ws.cell(row=r, column=STANDINGS_COLS[key]).value or 0
            w, otw, otl, l = get("w"), get("otw"), get("otl"), get("l")
            rows.append({
                "team": str(team).strip(),
                "w": int(w), "otw": int(otw), "otl": int(otl), "l": int(l),
                "points": int(get("points")), "gd": int(get("gd")),
                "gf": int(get("gf")), "ga": int(get("ga")),
                "gp": int(w + otw + otl + l),
            })
            r += 1
        return rows

    # First title is always the league-wide table.
    standings = read_table(title_rows[0][0])

    divisions = None
    if len(title_rows) > 1:
        divisions = {}
        for title_row, title_text in title_rows[1:]:
            # the sheet spells it "Divison" in some tabs and "Division" in others
            label = re.sub(r"\s*Divis(?:i)?on Standings\s*", "", title_text, flags=re.I).strip()
            teams_in_div = [row["team"] for row in read_table(title_row)]
            divisions[label] = teams_in_div

    return standings, divisions


def parse_schedule(ws):
    """Returns (completed, remaining, last_completed_date). Rows are assumed to be in
    chronological order (true of every mastersheet seen so far), so the last completed
    row encountered is the most recent game actually recorded in this file -- which is
    NOT necessarily "today". A file that's a few days stale will honestly report an
    older date here instead of us guessing "now" and being wrong."""
    completed, remaining = [], []
    last_completed_date = None
    for row in range(3, ws.max_row + 1):
        away = ws.cell(row=row, column=SCHED_COLS["away"]).value
        home = ws.cell(row=row, column=SCHED_COLS["home"]).value
        if not away or not home:
            continue
        d = ws.cell(row=row, column=SCHED_COLS["date"]).value
        game_date = str(d).strip() if d else None
        a_score = ws.cell(row=row, column=SCHED_COLS["away_score"]).value
        h_score = ws.cell(row=row, column=SCHED_COLS["home_score"]).value
        if a_score is not None and h_score is not None:
            ot_flag = ws.cell(row=row, column=SCHED_COLS["ot"]).value
            completed.append({
                "date": game_date, "away": str(away).strip(), "home": str(home).strip(),
                "away_score": float(a_score), "home_score": float(h_score),
                "ot": "OT" if isinstance(ot_flag, str) and ot_flag.strip().upper() == "OT" else None,
            })
            if game_date:
                last_completed_date = game_date
        else:
            remaining.append({"date": game_date, "away": str(away).strip(), "home": str(home).strip()})
    return completed, remaining, last_completed_date


def parse_roster(ws):
    totals = {}
    for row in range(3, ws.max_row + 1):
        team = ws.cell(row=row, column=ROSTER_COLS["team"]).value
        if not team or str(team).strip() in ("Free Agent", "Free Agents"):
            continue
        team = str(team).strip()
        t = totals.setdefault(team, {"G": 0.0, "GA": 0.0, "TK": 0.0, "GV": 0.0})
        t["G"] += ws.cell(row=row, column=ROSTER_COLS["g"]).value or 0
        t["GA"] += ws.cell(row=row, column=ROSTER_COLS["ga"]).value or 0
        t["TK"] += ws.cell(row=row, column=ROSTER_COLS["tk"]).value or 0
        t["GV"] += ws.cell(row=row, column=ROSTER_COLS["gv"]).value or 0
    return totals


def build_league(xlsx_path, meta, league_key):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    standings, divisions = parse_standings(wb["StandingsSchedule"])
    completed, remaining, last_completed_date = parse_schedule(wb["Backend Schedule"])
    roster = parse_roster(wb["Backend Players"])

    # Sanity check: total games played implied by the standings should roughly match
    # the completed-game count from the schedule tab. A big gap means this workbook's
    # Backend Schedule tab is behind its own StandingsSchedule tab (stale export).
    gp_total = sum(row["gp"] for row in standings)
    sched_game_participations = len(completed) * 2
    if gp_total - sched_game_participations >= len(standings):  # ~1+ game/team missing
        print(f"  WARNING [{league_key}] StandingsSchedule implies {gp_total} total team-games "
              f"played, but Backend Schedule only has {len(completed)} completed games "
              f"({sched_game_participations} team-games) -- this workbook looks stale/out of "
              f"sync with itself. Re-export a fresh copy of this mastersheet.", file=sys.stderr)

    colors = meta.get("colors", {})
    shorts = meta.get("short", {})
    missing = []
    palette_i = 0
    for row in standings:
        name = row["team"]
        if name not in colors:
            colors[name] = FALLBACK_PALETTE[palette_i % len(FALLBACK_PALETTE)]
            palette_i += 1
            missing.append(name)
        if name not in shorts:
            shorts[name] = name.split(" ")[0]
            if name not in missing:
                missing.append(name)
        row["color"] = colors[name]
        row["short"] = shorts[name]
        if name not in roster:
            print(f"  WARNING [{league_key}] no roster rows found for '{name}' -- "
                  f"treating G/GA/TK/GV as 0", file=sys.stderr)
            roster[name] = {"G": 0, "GA": 0, "TK": 0, "GV": 0}

    for name in missing:
        print(f"  WARNING [{league_key}] '{name}' is missing from team_meta.json "
              f"(colors/short) -- add it there so this stops using a placeholder",
              file=sys.stderr)

    ot_games = sum(1 for g in completed if g["ot"] == "OT")
    ot_rate = (ot_games / len(completed)) if completed else 0.15

    if last_completed_date:
        as_of = f"Standings and schedule through games of {last_completed_date} (per this mastersheet export)"
    else:
        today = date.today()
        as_of = f"Standings and schedule current as of {today.strftime('%B')} {today.day}, {today.year}"

    return {
        "standings": standings,
        "remaining": remaining,
        "divisions": divisions,
        "colors": colors,
        "short": shorts,
        "roster": {name: {"G": vals["G"], "GA": vals["GA"], "TK": vals["TK"], "GV": vals["GV"]}
                   for name, vals in roster.items() if name in {r["team"] for r in standings}},
        "otRate": ot_rate,
        "asOf": as_of,
        "playoffSpots": meta["playoffSpots"],
        "totalTeams": len(standings),
    }


def patch_html(html_path, leagues):
    text = Path(html_path).read_text(encoding="utf-8")
    lines = text.split("\n")
    json_str = json.dumps(leagues, ensure_ascii=False)
    for i, line in enumerate(lines):
        if line.strip().startswith("const LEAGUES = "):
            indent = line[:len(line) - len(line.lstrip())]
            lines[i] = f"{indent}const LEAGUES = {json_str};"
            break
    else:
        raise ValueError("Could not find 'const LEAGUES = ' line in the HTML file")

    backup = Path(html_path).with_suffix(".bak.html")
    backup.write_text(text, encoding="utf-8")
    Path(html_path).write_text("\n".join(lines), encoding="utf-8")
    print(f"Backed up previous version to {backup}")
    print(f"Wrote updated data into {html_path}")


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--premier", required=True,
                    help="Path to the Premier mastersheet .xlsx, or its Google Sheets URL")
    p.add_argument("--challenger", required=True,
                    help="Path to the Challenger mastersheet .xlsx, or its Google Sheets URL")
    p.add_argument("--html", required=True, help="Path to the playoff-odds HTML file to update in place")
    p.add_argument("--meta", default=str(Path(__file__).parent / "team_meta.json"),
                   help="Path to team_meta.json (default: next to this script)")
    args = p.parse_args()

    meta = json.loads(Path(args.meta).read_text(encoding="utf-8"))

    challenger_xlsx = resolve_xlsx(args.challenger, "Challenger")
    premier_xlsx = resolve_xlsx(args.premier, "Premier")

    print("Reading Challenger mastersheet...")
    challenger = build_league(challenger_xlsx, meta["challenger"], "challenger")
    print(f"  {len(challenger['standings'])} teams, "
          f"{len(challenger['remaining'])} games remaining, "
          f"OT rate {challenger['otRate']:.1%}")
    print(f"  {challenger['asOf']}")

    print("Reading Premier mastersheet...")
    premier = build_league(premier_xlsx, meta["premier"], "premier")
    print(f"  {len(premier['standings'])} teams, "
          f"{len(premier['remaining'])} games remaining, "
          f"OT rate {premier['otRate']:.1%}")
    print(f"  {premier['asOf']}")

    # Persist any auto-assigned fallback colors/shorts back so re-runs are stable
    # and so you notice what needs a real entry in team_meta.json.
    meta["challenger"]["colors"] = challenger["colors"]
    meta["challenger"]["short"] = challenger["short"]
    meta["premier"]["colors"] = premier["colors"]
    meta["premier"]["short"] = premier["short"]
    Path(args.meta).write_text(json.dumps(meta, indent=2, ensure_ascii=False), encoding="utf-8")

    patch_html(args.html, {"challenger": challenger, "premier": premier})
    print("Done. Open the HTML file and hit Run Simulation to confirm it looks right, "
          "then redeploy (e.g. drag the file into Netlify, or push it if the site is on git).")


if __name__ == "__main__":
    main()
